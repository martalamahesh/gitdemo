from selenium import webdriver
from selenium.webdriver.common.by import By
import time


from openpyxl import *
from datetime import datetime

driver = webdriver.Chrome()

#you can create one sample1.xlsx file in your project the below line just loads already crated workbook
wb = load_workbook('sample1.xlsx')

sheet = wb.active
# inputs
# i hardcoded date values
# if you want give at run time guse input fnction
# from_date= input("entere the form ")
#   to_date=input("entere the to_date ")
from_date = "15-08-2023"
to_date = "27-09-2023"
a = "https://www.centralbank.ie/news-media/press-releases?cat=-1&from={aa}&to={bb}".format(aa=from_date, bb=to_date)
# opening the browser

driver.get(a)
time.sleep(2)
driver.find_element(By.XPATH, "//*[@id='onetrust-accept-btn-handler']").click()
time.sleep(10)
elements = driver.find_elements(By.XPATH, "//*[@class='spotlight fixed-lines spotlight-wide new-story']")
count = len(elements)
# checking articles present or not and also no more than 10 articles
end = count + 1
pag = driver.find_elements(By.XPATH, "//ul[@class='pagination']/li")
p_count = len(pag)
r_value = 1
if count == 1:
    zero = driver.find_element(By.XPATH, "//*[@class='spotlight fixed-lines spotlight-wide new-story']/p").text
    zzz = "0 News Returned"
    if zero == zzz:
        print("no records to in the page")
elif 0 < count < 11 and p_count == 0:

    for i in range(1, end):
        # for j in range(1, 5):
        # speaker_name=
        speaker_xpath = "//div[@class='spotlight fixed-lines spotlight-wide new-story'][{fsp}]/img".format(fsp=i)
        attribute_name = "title"
        sheet.cell(row=i, column=1).value = driver.find_element(By.XPATH, speaker_xpath).get_attribute(
            attribute_name)
        # speach_date=
        date_xpath = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fda}]/div/span[1]".format(fda=i)
        sheet.cell(row=i, column=2).value = driver.find_element(By.XPATH, date_xpath).text
        # speach_title=
        title_xpath = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fti}]/div/a//*[@class='spotlight-title']".format(
            fti=i)
        sheet.cell(row=i, column=3).value = driver.find_element(By.XPATH, title_xpath).text
        # speach_url=
        attribute_name_h = "href"
        speach_url = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fsu}]/div/a".format(fsu=i)
        sheet.cell(row=i, column=4).value = driver.find_element(By.XPATH, speach_url).get_attribute(attribute_name_h)
        driver.find_element(By.XPATH, speach_url).click()
        time.sleep(10)

        # speach_content=
        sheet.cell(row=i, column=5).value = driver.find_element(By.XPATH,
                                                                "//div[@data-placeholder-label='News Article']/article").text
        driver.back()
elif p_count > 0:
    # i looped through pages based on neext page enable or disable
    try:
        f_elements = driver.find_elements(By.XPATH, "//*[@class='spotlight fixed-lines spotlight-wide new-story']")
        f_count = len(f_elements)
        # next_page = driver.find_element(By.XPATH, "//ul[@class='pagination']//li[@class='next-page']/a")

        while driver.find_element(By.XPATH, "//ul[@class='pagination']//li[@class='next-page']/a").is_enabled():
            elements = driver.find_elements(By.XPATH, "//*[@class='spotlight fixed-lines spotlight-wide new-story']")
            pg = len(elements)
            print(pg)
            f_pg = pg + 1
            print("page_records", f_pg)

            for i in range(1, f_pg):
                # print(r_value)
                # speaker_name=
                speaker_xpath = "//div[@class='spotlight fixed-lines spotlight-wide new-story'][{fsp}]/img".format(
                    fsp=i)
                attribute_name = "title"
                sheet.cell(row=r_value, column=1).value = driver.find_element(By.XPATH, speaker_xpath).get_attribute(
                    attribute_name)
                # speach_date=
                date_xpath = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fda}]/div/span[1]".format(
                    fda=i)
                sheet.cell(row=r_value, column=2).value = driver.find_element(By.XPATH, date_xpath).text
                # speach_title=
                speaker_xpath = "//div[@class='spotlight fixed-lines spotlight-wide new-story'][{fsp}]/img".format(
                    fsp=i)
                attribute_name = "title"
                sheet.cell(row=r_value, column=1).value = driver.find_element(By.XPATH, speaker_xpath).get_attribute(
                    attribute_name)
                # speach_url=
                attribute_name_h = "href"
                speach_url = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fsu}]/div/a".format(fsu=i)
                sheet.cell(row=r_value, column=4).value = driver.find_element(By.XPATH, speach_url).get_attribute(
                    attribute_name_h)
                driver.find_element(By.XPATH, speach_url).click()
                time.sleep(10)
                # speach_content=
                sheet.cell(row=r_value, column=5).value = driver.find_element(By.XPATH,
                                                                              "//div[@data-placeholder-label='News Article']/article").text
                driver.back()

                r_value = r_value + 1
                time.sleep(10)
            # time.sleep(100)
            # print("r_value is",r_value )

            driver.find_element(By.XPATH, "//ul[@class='pagination']//li[contains(@class,'next-page')]/a").click()

            time.sleep(10)



    except:
        # Handle the StaleElementReferenceException here

        print("r_value is", r_value)
        elements = driver.find_elements(By.XPATH, "//*[@class='spotlight fixed-lines spotlight-wide new-story']")
        pg = len(elements)
        print(pg)
        f_pg = pg + 1
        print("page_records", f_pg)
        for i in range(1, f_pg):
            # print(r_value)

            speaker_xpath = "//div[@class='spotlight fixed-lines spotlight-wide new-story'][{fsp}]/img".format(
                fsp=i)
            attribute_name = "title"
            sheet.cell(row=r_value, column=1).value = driver.find_element(By.XPATH, speaker_xpath).get_attribute(
                attribute_name)
            # speach_date=
            date_xpath = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fda}]/div/span[1]".format(
                fda=i)
            sheet.cell(row=r_value, column=2).value = driver.find_element(By.XPATH, date_xpath).text
            # speach_title=
            speaker_xpath = "//div[@class='spotlight fixed-lines spotlight-wide new-story'][{fsp}]/img".format(
                fsp=i)
            attribute_name = "title"
            sheet.cell(row=r_value, column=1).value = driver.find_element(By.XPATH, speaker_xpath).get_attribute(
                attribute_name)
            # speach_url=
            attribute_name_h = "href"
            speach_url = "//*[@class='spotlight fixed-lines spotlight-wide new-story'][{fsu}]/div/a".format(fsu=i)
            sheet.cell(row=r_value, column=4).value = driver.find_element(By.XPATH, speach_url).get_attribute(
                attribute_name_h)
            driver.find_element(By.XPATH, speach_url).click()
            time.sleep(10)
            # speach_content=
            sheet.cell(row=r_value, column=5).value = driver.find_element(By.XPATH,
                                                                          "//div[@data-placeholder-label='News Article']/article").text
            driver.back()

wb.save("sample1.xlsx")
wb.close()

# print("mahesh")
